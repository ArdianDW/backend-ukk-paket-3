import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from API.models import inventaris
from API.serializers.inventaris_serializer import InventarisSerializer
from django.http import HttpResponse
from django.utils import timezone
import locale

class ExportLaporanView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        laporan_type = request.query_params.get('type', 'all')

        locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')

        if laporan_type == 'all':
            inventaris_list = inventaris.objects.all()
            title = 'Laporan Semua Barang Inventaris'
            filename = 'laporan_semua_barang_inventaris.xlsx'
        elif laporan_type == 'recent':
            now = timezone.now()
            inventaris_list = inventaris.objects.filter(
                tanggal_register__year=now.year,
                tanggal_register__month=now.month
            )
            bulan = now.strftime("%B")
            tahun = now.year
            title = f'Laporan Barang Baru Masuk Bulan {bulan} Tahun {tahun}'
            filename = f'laporan_barang_masuk_{bulan}_{tahun}.xlsx'
        elif laporan_type == 'rusak':
            inventaris_list = inventaris.objects.filter(kondisi='rusak')
            title = 'Laporan Barang Rusak'
            filename = 'laporan_barang_rusak.xlsx'
        elif laporan_type == 'hilang':
            inventaris_list = inventaris.objects.filter(kondisi='hilang')
            title = 'Laporan Inventaris Hilang'
            filename = 'laporan_inventaris_hilang.xlsx'
        else:
            return Response({'error': 'Invalid report type'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = InventarisSerializer(inventaris_list, many=True)
        data = serializer.data

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Laporan Inventaris'

        sheet.merge_cells('A1:I1')
        title_cell = sheet['A1']
        title_cell.value = title
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        sheet['A2'] = f'Diekspor oleh: {request.user.nama_petugas}'
        sheet['A2'].font = Font(italic=True)

        headers = ['No', 'Nama', 'Kondisi', 'Keterangan', 'Jumlah', 'Tanggal Register', 'Kode Inventaris', 'Nama Jenis', 'Nama Ruang']
        sheet.append(headers)

        for index, item in enumerate(data, start=1):
            sheet.append([
                index,  
                item['nama'],
                item['kondisi'],
                item['keterangan'],
                item['jumlah'],
                item['tanggal_register'],
                item['kode_inventaris'],
                item['nama_jenis'],
                item['nama_ruang']
            ])

        for column_cells in sheet.iter_cols(min_row=3, max_row=sheet.max_row, max_col=sheet.max_column):
            max_length = 0
            column = column_cells[0].column_letter  
            for cell in column_cells:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        virtual_workbook = BytesIO()
        workbook.save(virtual_workbook)
        virtual_workbook.seek(0)

        response = HttpResponse(content=virtual_workbook.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response 