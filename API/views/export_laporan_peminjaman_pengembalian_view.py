import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from API.models import RiwayatPeminjaman
from API.serializers.laporan_peminjaman_pengembalian_serializer import LaporanPeminjamanPengembalianSerializer
from django.http import HttpResponse
from django.utils import timezone
import locale

class ExportLaporanPeminjamanPengembalianView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        laporan_type = request.query_params.get('type', 'peminjaman')
        now = timezone.now()

        # Set locale to Indonesian
        locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')

        bulan = now.strftime("%B")
        tahun = now.year

        if laporan_type == 'peminjaman':
            riwayat_list = RiwayatPeminjaman.objects.filter(
                peminjaman__tanggal_pinjam__year=now.year,
                peminjaman__tanggal_pinjam__month=now.month
            )
            title = f'Laporan Peminjaman Bulan {bulan} Tahun {tahun}'
            filename = f'laporan_peminjaman_{bulan}_{tahun}.xlsx'
        elif laporan_type == 'pengembalian':
            riwayat_list = RiwayatPeminjaman.objects.filter(
                peminjaman__tanggal_kembali__year=now.year,
                peminjaman__tanggal_kembali__month=now.month
            )
            title = f'Laporan Pengembalian Bulan {bulan} Tahun {tahun}'
            filename = f'laporan_pengembalian_{bulan}_{tahun}.xlsx'
        else:
            return Response({'error': 'Invalid report type'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = LaporanPeminjamanPengembalianSerializer(riwayat_list, many=True, context={'request': request})
        data = serializer.data

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Laporan'

        sheet.merge_cells('A1:E1')
        title_cell = sheet['A1']
        title_cell.value = title
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        sheet['A2'] = f'Diekspor oleh: {request.user.nama_petugas}'
        sheet['A2'].font = Font(italic=True)

        headers = ['No', 'Nama Peminjam', 'Nama Barang', 'Jumlah Total', 'Tanggal']
        sheet.append(headers)

        for index, item in enumerate(data, start=1):
            for barang in item['barang_dipinjam']:
                sheet.append([
                    index,
                    item['nama_peminjam'],
                    barang['nama_barang'],
                    barang['jumlah'],
                    item['tanggal']
                ])

        for column_cells in sheet.iter_cols(min_row=3, max_row=sheet.max_row, max_col=sheet.max_column):
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        virtual_workbook = BytesIO()
        workbook.save(virtual_workbook)
        virtual_workbook.seek(0)

        response = HttpResponse(content=virtual_workbook.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response 